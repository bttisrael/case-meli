"""Regression tests for the optimization implemented in ``main.ipynb``.

The project intentionally keeps the business analysis in a notebook.  The
session fixture below executes the notebook's actual core model cells in their
original order, so these tests cannot silently drift to a second copy of the
optimization logic.
"""

from __future__ import annotations

import json
import os
from pathlib import Path

import pyomo.environ as pyo
import pytest
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
NOTEBOOK_PATH = PROJECT_ROOT / "main.ipynb"
INPUT_WORKBOOK = PROJECT_ROOT / "data" / "BicyclesRelocationData.xlsx"
OUTPUT_WORKBOOK = (
    PROJECT_ROOT / "results" / "BicyclesRelocationData_Optimized.xlsx"
)
OUTPUT_SHEET_NAME = "OptimalRelocationPlan"

CORE_MODEL_CELL_IDS = (
    "010a5a97",  # imports
    "1ca59ece",  # input and output paths
    "4f52432c",  # source worksheet discovery
    "1d7c7c91",  # workbook loading
    "full-input-code",
    "full-parameters-code",
    "full-variables-code",
    "full-constraints-code",
    "full-objective-code",
    "full-solve-code",
    "full-extract-code",
    "full-validation-code",
)

EXPECTED_CATEGORIES = [
    "Child",
    "Adult",
    "Electric",
    "Racing",
    "Mountain",
    "Tricycle",
]
EXPECTED_SURPLUS = {
    "Child": 272,
    "Adult": 270,
    "Electric": 279,
    "Racing": 267,
    "Mountain": 282,
    "Tricycle": 279,
}
EXPECTED_SPACE = {
    "Child": 1.0,
    "Adult": 1.5,
    "Electric": 1.5,
    "Racing": 1.7,
    "Mountain": 1.7,
    "Tricycle": 4.0,
}
EXPECTED_HORIZON = {
    "Child": 80,
    "Adult": 53,
    "Electric": 53,
    "Racing": 47,
    "Mountain": 47,
    "Tricycle": 20,
}
EXPECTED_POSITIVE_QUANTITIES = {
    (1, "Child"): 1,
    (3, "Child"): 69,
    (3, "Adult"): 1,
    (4, "Child"): 7,
    (6, "Adult"): 1,
}
EXPECTED_POSITIVE_PROFITS = {
    (1, "Child"): 45.4865,
    (3, "Child"): 3769.8666,
    (3, "Adult"): 70.1450,
    (4, "Child"): 392.0342,
    (6, "Adult"): 81.2511,
}


def _source_text(cell: dict) -> str:
    source = cell.get("source", "")
    return "".join(source) if isinstance(source, list) else source


@pytest.fixture(scope="session")
def model_namespace() -> dict:
    """Run only the data-driven MILP cells from the production notebook."""

    notebook = json.loads(NOTEBOOK_PATH.read_text(encoding="utf-8"))
    code_by_id = {
        cell.get("id"): _source_text(cell)
        for cell in notebook["cells"]
        if cell.get("cell_type") == "code"
    }
    missing_ids = [cell_id for cell_id in CORE_MODEL_CELL_IDS if cell_id not in code_by_id]

    if missing_ids:
        pytest.fail(f"Core model cells are missing from main.ipynb: {missing_ids}")

    namespace = {
        "__name__": "main_notebook_under_test",
        "__file__": str(NOTEBOOK_PATH),
        "display": lambda *_args, **_kwargs: None,
    }
    previous_directory = Path.cwd()

    try:
        os.chdir(PROJECT_ROOT)

        for cell_id in CORE_MODEL_CELL_IDS:
            source = code_by_id[cell_id]
            exec(compile(source, f"main.ipynb::{cell_id}", "exec"), namespace)
    finally:
        os.chdir(previous_directory)

    return namespace


def test_input_data_matches_the_case_contract(model_namespace):
    assert model_namespace["full_areas"] == list(range(1, 8))
    assert model_namespace["full_categories"] == EXPECTED_CATEGORIES
    assert model_namespace["full_surplus"] == EXPECTED_SURPLUS
    assert model_namespace["full_space"] == pytest.approx(EXPECTED_SPACE)
    assert model_namespace["full_horizon"] == EXPECTED_HORIZON
    assert sum(model_namespace["full_observed_length"].values()) == 3321
    assert all(value >= 0 for value in model_namespace["full_profit"].values())


def test_model_has_the_expected_binary_structure(model_namespace):
    model = model_namespace["full_model"]

    assert len(model.I) == 2100
    assert len(model.I_SEQUENCE) == 2058
    assert len(model.y) == 2100
    assert all(model.y[index].is_binary() for index in model.I)
    assert len(model.availability_constraint) == len(EXPECTED_CATEGORIES)
    assert len(model.sequence_constraint) == 2058
    assert model.capacity_constraint.active
    assert model.objective.sense == pyo.maximize


def test_solver_returns_the_known_unique_optimum(model_namespace):
    assert (
        model_namespace["full_termination"]
        == pyo.TerminationCondition.optimal
    )
    assert model_namespace["full_objective_value"] == pytest.approx(
        4358.7834,
        abs=1e-6,
    )
    assert model_namespace["full_bicycles_relocated"] == 79
    assert model_namespace["full_capacity_used"] == pytest.approx(80.0)
    assert model_namespace["full_capacity_remaining"] == pytest.approx(0.0)

    positive_plan = model_namespace["full_positive_plan"]
    quantities = positive_plan.set_index(["Area", "Category"])[
        "Bicycles Relocated"
    ].astype(int).to_dict()
    profits = positive_plan.set_index(["Area", "Category"])[
        "Expected Profit"
    ].to_dict()

    assert quantities == EXPECTED_POSITIVE_QUANTITIES
    assert profits == pytest.approx(EXPECTED_POSITIVE_PROFITS, abs=1e-6)


def test_solution_is_binary_and_respects_every_prefix(model_namespace):
    selected = model_namespace["full_selected"]

    assert set(selected.values()) <= {0, 1}
    assert all(
        selected[area, category, position]
        <= selected[area, category, position - 1]
        for area, category, position in model_namespace["full_model"].I_SEQUENCE
    )
    assert model_namespace["full_missing_tail_selected"] == 0


def test_solution_respects_global_availability_and_capacity(model_namespace):
    selected = model_namespace["full_selected"]
    category_totals = {
        category: sum(
            selected[area, indexed_category, position]
            for area, indexed_category, position in selected
            if indexed_category == category
        )
        for category in model_namespace["full_categories"]
    }

    assert category_totals == {
        "Child": 77,
        "Adult": 2,
        "Electric": 0,
        "Racing": 0,
        "Mountain": 0,
        "Tricycle": 0,
    }
    assert all(
        category_totals[category] <= model_namespace["full_surplus"][category]
        for category in model_namespace["full_categories"]
    )

    capacity_used = sum(
        model_namespace["full_space"][category] * chosen
        for (_area, category, _position), chosen in selected.items()
    )
    assert capacity_used == pytest.approx(80.0)
    assert capacity_used <= model_namespace["FULL_TRUCK_CAPACITY"]


def test_objective_reconciles_with_selected_marginal_profits(model_namespace):
    recomputed_profit = sum(
        model_namespace["full_profit"][index] * chosen
        for index, chosen in model_namespace["full_selected"].items()
    )
    plan_profit = model_namespace["full_plan"]["Expected Profit"].sum()

    assert recomputed_profit == pytest.approx(4358.7834, abs=1e-6)
    assert plan_profit == pytest.approx(recomputed_profit, abs=1e-6)
    assert model_namespace["full_validation"]["Passed"].all()


def test_export_preserves_source_tabs_and_reconciles_the_solution(model_namespace):
    source_workbook = load_workbook(
        INPUT_WORKBOOK,
        read_only=True,
        data_only=False,
    )
    output_workbook = load_workbook(
        OUTPUT_WORKBOOK,
        read_only=True,
        data_only=False,
    )

    try:
        assert output_workbook.sheetnames == [
            OUTPUT_SHEET_NAME,
            *source_workbook.sheetnames,
        ]

        for sheet_name in source_workbook.sheetnames:
            assert tuple(source_workbook[sheet_name].values) == tuple(
                output_workbook[sheet_name].values
            )

        plan_sheet = output_workbook[OUTPUT_SHEET_NAME]
        assert plan_sheet["B3"].value == "Exact MILP (HiGHS)"
        assert plan_sheet["B4"].value == "Optimal"
        assert plan_sheet["B5"].value == pytest.approx(4358.7834, abs=1e-6)
        assert plan_sheet["B6"].value == 79
        assert plan_sheet["B7"].value == pytest.approx(80.0)
        assert plan_sheet["B8"].value == pytest.approx(0.0)

        exported_quantities = {}
        exported_profits = {}

        for row in plan_sheet.iter_rows(
            min_row=12,
            max_col=6,
            values_only=True,
        ):
            if row[0] == "TOTAL":
                assert row[2] == 79
                assert row[4] == pytest.approx(80.0)
                assert row[5] == pytest.approx(4358.7834, abs=1e-6)
                break

            if row[0] is None:
                continue

            key = (int(row[0]), row[1])
            exported_quantities[key] = int(row[2])
            exported_profits[key] = float(row[5])

        assert exported_quantities == EXPECTED_POSITIVE_QUANTITIES
        assert exported_profits == pytest.approx(EXPECTED_POSITIVE_PROFITS, abs=1e-6)
        assert all(plan_sheet[f"J{row}"].value is True for row in range(23, 30))
    finally:
        source_workbook.close()
        output_workbook.close()
